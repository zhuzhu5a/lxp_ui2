import xlrd
import openpyxl
# class excelUtil():
#     def __init__(self,excelPath,sheetName='Sheet1'):
#         self.data =xlrd.open_workbook(excelPath)
#         self.table = self.data.sheet_by_name(sheetName)
#         self.keys = self.table.row_values(0)
#         self.rowNum = self.table.nrows
#         self.colNum = self.table.ncols
#     def dict_data(self):
#             if self.rowNum<=1:
#                 print('总行数小于1')
#             else:
#                 r=[]
#                 j= 1
#                 for i in range(self.rowNum-1):
#                     s= {}
#                     values =self.table.row_values(j)
#                     for x in range(self.colNum):
#                         s[self.keys[x]]=values[x]
#                     r.append(s)
#                     j+=1
#                 return r
# if __name__ == '__main__':
#     filepath = r'C:\Users\91621\PycharmProjects\lxp项目\data\login.xls'
#     data=excelUtil(filepath)
#     print(data.dict_data())
# def test_data(path):
#     with open(path,encoding='utf-8',mode='r') as f:
#         data=f.readlines()
#         mydata= []
#         for line in data:
#             mydata.append(line.strip())
#             f.close()
#             return mydata
class GetData():
    def getExcel(self,filepath,sheet):
        wk=openpyxl.load_workbook(filepath)
        sheet=wk[sheet]
        # print(value)
        datalist=[]
        for row in range (2,sheet.max_row+1):
            rowdata={}
            for col in range(1,sheet.max_column+1):
                key=sheet.cell(1,col).value
                value=sheet.cell(row,col).value
                rowdata[key]=value
            datalist.append(rowdata)
        return datalist
    def getExcel_row(self,filepath,sheet):
        wk = openpyxl.load_workbook(filepath)
        sheet = wk[sheet]
        row=sheet.max_row
        return row

if __name__ == '__main__':
    data=GetData().getExcel()
    print(data)

